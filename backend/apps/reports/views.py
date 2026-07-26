import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from rest_framework import status, permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from apps.donations.models import Donation
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

class ExportExcelReportView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        user = request.user
        token = request.query_params.get('token')
        if token and (not user or not user.is_authenticated):
            from rest_framework_simplejwt.tokens import AccessToken
            from django.contrib.auth import get_user_model
            User = get_user_model()
            try:
                validated_token = AccessToken(token)
                user = User.objects.get(id=validated_token['user_id'])
            except Exception:
                user = None

        if not user or not user.is_authenticated or not user.is_org_admin():
            return Response({'error': 'Permission denied. Admin access required.'}, status=status.HTTP_403_FORBIDDEN)

        report_type = request.query_params.get('type', 'all')
        donations = Donation.objects.filter(payment_status='SUCCESS').order_by('-created_at')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RAHMA Donation Report"

        # Headers & Styling
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )

        headers = ["Receipt #", "Donation Date", "Donor Name", "Email", "Phone", "Campaign", "Amount (INR)", "Payment Method", "Transaction ID"]
        ws.append(headers)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center

        for d in donations:
            row = [
                d.receipt_number,
                d.created_at.strftime('%Y-%m-%d %H:%M'),
                d.donor_name,
                d.donor_email,
                d.donor_phone or 'N/A',
                d.campaign.title if d.campaign else 'General Donation',
                float(d.amount),
                d.get_payment_method_display(),
                d.razorpay_payment_id or d.donation_number
            ]
            ws.append(row)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="RAHMA_Donations_Report.xlsx"'
        return response

class ExportPDFReportView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        user = request.user
        token = request.query_params.get('token')
        if token and (not user or not user.is_authenticated):
            from rest_framework_simplejwt.tokens import AccessToken
            from django.contrib.auth import get_user_model
            User = get_user_model()
            try:
                validated_token = AccessToken(token)
                user = User.objects.get(id=validated_token['user_id'])
            except Exception:
                user = None

        if not user or not user.is_authenticated or not user.is_org_admin():
            return Response({'error': 'Permission denied. Admin access required.'}, status=status.HTTP_403_FORBIDDEN)

        donations = Donation.objects.filter(payment_status='SUCCESS').order_by('-created_at')[:50]

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        story = []
        styles = getSampleStyleSheet()

        EMERALD_GREEN = colors.HexColor('#064e3b')
        LIGHT_GREEN = colors.HexColor('#ecfdf5')

        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=20, textColor=EMERALD_GREEN)
        story.append(Paragraph("RAHMA Financial Donation Summary Report", title_style))
        story.append(HRFlowable(width="100%", thickness=1.5, color=EMERALD_GREEN, spaceBefore=5, spaceAfter=15))

        table_data = [["Receipt #", "Date", "Donor Name", "Campaign", "Amount", "Method"]]
        header_style = ParagraphStyle('HeaderStyle', fontName='Helvetica-Bold', fontSize=9, textColor=colors.white)
        cell_style = ParagraphStyle('CellStyle', fontName='Helvetica', fontSize=8)

        for d in donations:
            table_data.append([
                Paragraph(d.receipt_number, cell_style),
                Paragraph(d.created_at.strftime('%Y-%m-%d'), cell_style),
                Paragraph(d.donor_name, cell_style),
                Paragraph(d.campaign.title[:20] if d.campaign else 'General', cell_style),
                Paragraph(f"₹{d.amount:,.2f}", cell_style),
                Paragraph(d.get_payment_method_display(), cell_style),
            ])

        report_table = Table(table_data, colWidths=[90, 70, 130, 110, 75, 60])
        report_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), EMERALD_GREEN),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d1d5db')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT_GREEN]),
            ('PADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(report_table)

        doc.build(story)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="RAHMA_Financial_Report.pdf"'
        return response
